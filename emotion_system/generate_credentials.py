"""
Auto-Generate Credentials with Random Passwords
================================================
Reads StudentPicsDataset.csv and creates login credentials
for every student using their Student ID as username
and a unique random password for each student.
"""

import sys, os, json, random, string, pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from gui.data_store import DataStore


def random_password(length=8):
    """Generate a random password: letters + numbers."""
    chars = string.ascii_letters + string.digits
    return ''.join(random.choices(chars, k=length))


def generate_credentials(csv_path: str):
    if not os.path.exists(csv_path):
        print(f"❌ File not found: {csv_path}")
        return

    print(f"📖 Reading: {csv_path}\n")
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    print(f"Found {len(df)} students\n{'='*60}")

    store = DataStore()
    added = 0
    skipped = 0
    credentials_list = []

    for idx, row in df.iterrows():
        student_id = str(row['Student ID']).strip()
        name       = str(row['Student Name']).strip()

        username = student_id
        password = random_password()

        # Check if already exists
        existing = store.get_user(username)
        if existing:
            print(f"  ⚠️  [{idx+1}] {name} — already exists, skipping")
            skipped += 1
            continue

        # Find student in store
        student = next((s for s in store.students
                       if str(student_id) in s.get('email','') or
                       s.get('name','') == name), None)

        store.add_user({
            "username":   username,
            "password":   password,
            "role":       "student",
            "name":       name,
            "email":      f"{student_id}@university.edu",
            "student_id": student["id"] if student else "",
        })

        credentials_list.append({
            "student_id": student_id,
            "name":       name,
            "username":   username,
            "password":   password,
        })

        print(f"  ✅ [{idx+1}] {name} → {username} / {password}")
        added += 1

    # Save to store.json
    try:
        store_path = 'gui/store.json'
        existing_data = {}
        if os.path.exists(store_path):
            with open(store_path, 'r', encoding='utf-8') as f:
                existing_data = json.load(f)

        existing_data['users'] = [u for u in store.users
                                   if u['username'] not in
                                   ('admin','dr.smith','s001','parent1')]

        with open(store_path, 'w', encoding='utf-8') as f:
            json.dump(existing_data, f, indent=2, ensure_ascii=False)

        print(f"\n{'='*60}")
        print(f"✅ Done!")
        print(f"   Added: {added} credentials")
        print(f"   Skipped: {skipped}")
        print(f"{'='*60}")

    except Exception as e:
        print(f"❌ Save error: {e}")

    # Export credentials to Excel
    if credentials_list:
        export_csv  = 'student_credentials.csv'
        export_xlsx = 'student_credentials.xlsx'

        # Save as CSV
        with open(export_csv, 'w', encoding='utf-8-sig') as f:
            f.write("Student ID,Name,Username,Password\n")
            for c in credentials_list:
                f.write(f"{c['student_id']},{c['name']},{c['username']},{c['password']}\n")

        # Save as Excel
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Student Credentials"

            # Header
            headers = ["Student ID", "Name", "Username", "Password"]
            header_fill = PatternFill("solid", fgColor="1E3A6E")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Set column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            # Data rows
            for row_idx, c in enumerate(credentials_list, 2):
                ws.cell(row=row_idx, column=1, value=c['student_id'])
                ws.cell(row=row_idx, column=2, value=c['name'])
                ws.cell(row=row_idx, column=3, value=c['username'])
                ws.cell(row=row_idx, column=4, value=c['password'])

                # Alternate row colors
                if row_idx % 2 == 0:
                    fill = PatternFill("solid", fgColor="F0F4FF")
                    for col in range(1, 5):
                        ws.cell(row=row_idx, column=col).fill = fill

            wb.save(export_xlsx)
            print(f"\n📁 Credentials saved to Excel: {export_xlsx}")

        except Exception as e:
            print(f"\n📁 Credentials saved to CSV: {export_csv}")

        print(f"   Open the file in Excel to see all usernames and passwords!")
        print(f"\n💡 Each student has a UNIQUE random password!")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_credentials.py StudentPicsDataset.csv")
        sys.exit(1)

    generate_credentials(sys.argv[1])
